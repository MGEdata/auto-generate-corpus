# -*- coding: utf8 -*-

# import csv
import pandas as pd
from preprocessors import *

from openpyxl import load_workbook

#测试集
xlsx_file="result_dev.xlsx"
#训练集
# xlsx_file2="result675.xlsx"
# true_re="result_true.xlsx"

#将正确的标签读出来

#读取路径
# book = load_workbook(filename=true_re)
# #读取名字为Sheet1的表
# sheet = book.get_sheet_by_name("result")
# #用于存储数据的数组
# data= []
# row_num = 2
# while row_num <= 511:
#     #将表中第一列的1-1000行数据写入data数组中
#     data.append(sheet.cell(row=row_num, column=1).value)
#     row_num = row_num + 1

# print(len(data))
# print(data)

#df1 为datafram格式
## 以下将person1、person2格式转化成元组
df = pd.read_excel(xlsx_file, usecols=[0,1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,13,14,15,16,17,18,19,20])

person1_word_idx=df['person1_word_idx'].values
person2_word_idx=df['person2_word_idx'].values
# print(person1_word_idx)
person1_t=[]
person2_t=[]

for i in person1_word_idx:
    # print(i)
    temp = i.replace('(', '').replace(')', '')
    a = tuple([int(i) for i in temp.split(':')])
    person1_t.append(a)

for i in person2_word_idx:
    # print(i)
    temp = i.replace('(', '').replace(')', '')
    a = tuple([int(i) for i in temp.split(':')])
    person2_t.append(a)

df['person1_word_idx_t'] =person1_t
df['person2_word_idx_t'] =person2_t
df1=df.drop(['person1_word_idx','person2_word_idx'],axis=1)


df1.rename(columns={'person1_word_idx_t':'person1_word_idx', 'person2_word_idx_t':'person2_word_idx'}, inplace = True)


#换位置
person1= df1['person1_word_idx']
person2=df1['person2_word_idx']
df1.drop(labels=['person1_word_idx'], axis=1,inplace = True)
df1.drop(labels=['person2_word_idx'], axis=1,inplace = True)

df1.insert(0, 'person1_word_idx', person1)
df1.insert(1, 'person2_word_idx', person2)

df22=df1
# print(df1.columns.values.tolist())

#去掉tokes 中"'"和","; 并且将tokens改成列表格式。
# df22 = pd.read_excel(xlsx_file, usecols=[0,1, 2, 3, 4, 5, 6])
tokenList=[]
for row in df22.itertuples():
    # print(row)
    person1=getattr(row,'person1_word_idx')
    #返回对象的属性值
    tokens=getattr(row,'tokens')
    # print(tokens)
    newtokens = tokens.replace(']','').replace('[','')
    # print(newtokens)

    nt=newtokens.split(',');
    while '' in nt:
        nt.remove('')
    # print(type(nt))
    # print(nt)
    tokenList.append(nt)

#将tokens写到df中，以下都执行成功。
df22['tokens2'] = tokenList

# print(df22)
df33=df22

pd.options.display.max_columns = 100
pd.options.display.max_rows = 100




df33= df33.drop(['tokens'], axis=1)
df33.rename(columns={'tokens2':'tokens'}, inplace = True)
# print(df33)


#df33 需要的列格式转化成功
#调用preprocessors.py 中的处理方法
#
list_1=[]
for i in range(len(df33)):
    cand=df33.loc[i]
    # print(cand)
    # print(type(cand))
    cand2=get_text_between(cand)
    cand3=get_left_tokens(cand2)
    cand4 =get_right_first_tokens(cand3)
    cand5= get_right_second_tokens(cand4)
    cand6= get_right_third_tokens(cand5)
    cand7 = get_left_first_tokens(cand6)
    cand8 = get_left_second_tokens(cand7)
    cand9 = get_left_third_tokens(cand8)
    cand10 = get_right_tokens(cand9)
    list_1.append(cand10)
    # print(cand6)

# for i in list_1:
#     print(i)
#     print(type(i))



df44 = pd.DataFrame(list(list_1[i] for i in range(0, len(list_1))))

pd.set_option('max_colwidth',10000)


# for index, row in 44.iterrows():
#     print(row) # 输出每行的索引值
# print(df44)
# print(type(df44[0]))
# print(df44['person1_word_id'])
df44.to_pickle("./data2/dev_data.pkl")



    # s=df.iloc[i]
    # df1=get_text_between(s)
    # print(df1)












