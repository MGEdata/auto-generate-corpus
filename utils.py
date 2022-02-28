import os
import pickle
import subprocess
from typing import Tuple
import numpy as np

import pandas as pd
from openpyxl import load_workbook

#IS_TEST = os.environ.get("TRAVIS") == "true" or os.environ.get("IS_TEST") == "False"
IS_TEST="true"


def load_data() -> Tuple[
    Tuple[pd.DataFrame, np.ndarray], pd.DataFrame, Tuple[pd.DataFrame, np.ndarray]
]:
    """
    Returns:
        df_dev, Y_dev: Development set data points and 1D labels ndarray.
        df_train: Training set data points dataframe.
        df_test, Y_test: Test set data points dataframe and 1D labels ndarray.
    """
    # try:
    #     subprocess.run(["bash", "download_data.sh"], check=True, stderr=subprocess.PIPE)
    # except subprocess.CalledProcessError as e:
    #     print(e.stderr.decode())
    #     raise e

    # with open(os.path.join("./data2", "dev_data.pkl"), "rb") as f:
    #     # df_dev = pickle.load(f)
    #     Y_dev = pickle.load(f)
    true_re = "result_true_test.xlsx"
    book = load_workbook(filename=true_re)
    # 读取名字为Sheet1的表
    sheet = book.get_sheet_by_name("result")
    # 用于存储数据的数组
    Y_test = []
    row_num = 2
    while row_num <= 311:
        # 将表中第一列的1-1000行数据写入data数组中
        Y_test.append(sheet.cell(row=row_num, column=1).value)
        row_num = row_num + 1

    true_re1 = "result_true_dev.xlsx"
    book1 = load_workbook(filename=true_re1)
    # 读取名字为Sheet1的表
    sheet1 = book1.get_sheet_by_name("result")
    # 用于存储数据的数组
    Y_dev = []
    row_num = 2
    while row_num <= 201:
        # 将表中第一列的1-1000行数据写入data数组中
        Y_dev.append(sheet1.cell(row=row_num, column=1).value)
        row_num = row_num + 1


    with open(os.path.join("./data2", "train_data.pkl"), "rb") as f:
        df_train = pickle.load(f)
        if IS_TEST:
            # Reduce train set size to speed up travis.
            df_train = df_train.iloc[:2000]

    with open(os.path.join("./data2", "test_data.pkl"), "rb") as f:
        df_test = pickle.load(f)


    with open(os.path.join("./data2", "dev_data.pkl"), "rb") as f:
        df_dev = pickle.load(f)
        # Y_dev = pickle.load(f)


    # Convert labels to {0, 1} format from {-1, 1} format.
    if Y_dev !=None:
        Y_dev = (1 + np.array(Y_dev)) // 2
        Y_test = (1 + np.array(Y_test)) // 2

    #将dev和test设为none
    # df_dev=None
    # Y_dev=None
    # df_test=None
    # Y_test=None
    #将df_dev转化成np.array格式，
    return ((df_dev, Y_dev), df_train, (df_test, Y_test))


def get_n_epochs() -> int:
    return 10 if IS_TEST else 1